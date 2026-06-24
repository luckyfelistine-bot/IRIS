"""IRIS v7 Document Processing — Unlimited Files with Vector Search"""
import os
import re
import json
import tempfile
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
from config import config
from db import db

class DocumentProcessor:
    """
    Process unlimited documents with:
    - 50+ file type support
    - OCR for images
    - Text extraction and chunking
    - Vector embeddings for semantic search
    - Full-text search (FTS)
    - Document analysis and summarization
    """

    SUPPORTED_TYPES = {
        'text': ['.txt', '.md', '.rst', '.csv', '.json', '.xml', '.yaml', '.yml', '.log', '.ini', '.conf'],
        'code': ['.py', '.js', '.ts', '.jsx', '.tsx', '.html', '.css', '.scss', '.java', '.cpp', '.c', '.h', 
                 '.go', '.rs', '.rb', '.php', '.swift', '.kt', '.scala', '.r', '.m', '.sql', '.sh', '.bat',
                 '.vue', '.svelte', '.astro', '.php', '.pl', '.lua', '.dart', '.fs', '.fsx', '.clj', '.cljs',
                 '.erl', '.ex', '.exs', '.ml', '.mli', '.hs', '.lhs', '.idr', '.nim', '.cr', '.v', '.zig'],
        'document': ['.pdf', '.docx', '.doc', '.odt', '.rtf', '.epub', '.tex', '.odf'],
        'spreadsheet': ['.xlsx', '.xls', '.ods', '.csv', '.tsv'],
        'presentation': ['.pptx', '.ppt', '.odp'],
        'image': ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp', '.svg', '.ico', '.heic', '.raw'],
        'audio': ['.mp3', '.wav', '.ogg', '.flac', '.aac', '.m4a', '.wma'],
        'video': ['.mp4', '.avi', '.mkv', '.mov', '.wmv', '.flv', '.webm', '.m4v', '.3gp'],
        'archive': ['.zip', '.tar', '.gz', '.bz2', '.7z', '.rar', '.xz', '.lz', '.lzma']
    }

    def __init__(self):
        self.upload_dir = config.UPLOADS_DIR
        os.makedirs(self.upload_dir, exist_ok=True)
        self.vector_store = self._init_vector_store()

    def _init_vector_store(self):
        """Initialize in-memory vector store (upgrade to FAISS/Chroma for production)"""
        try:
            from sentence_transformers import SentenceTransformer
            self.encoder = SentenceTransformer('all-MiniLM-L6-v2')
            self.vectors = {}
            self.documents = {}
            return True
        except Exception as e:
            db.log("WARNING", "document_processor", f"Vector store init failed: {e}")
            self.encoder = None
            return False

    def get_file_type(self, filename: str) -> str:
        """Determine file type from extension"""
        ext = Path(filename).suffix.lower()
        for file_type, extensions in self.SUPPORTED_TYPES.items():
            if ext in extensions:
                return file_type
        return 'unknown'

    def process_upload(self, file_path: str, original_name: str, 
                      extract_images: bool = True, 
                      generate_summary: bool = True) -> Dict:
        """
        Process uploaded file:
        1. Save to storage
        2. Extract text/content
        3. Chunk for vector search
        4. Generate summary
        5. Store in database
        """
        try:
            file_type = self.get_file_type(original_name)
            file_size = os.path.getsize(file_path)

            # Save to permanent storage
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_name = re.sub(r'[^\w\-.]', '_', original_name)
            stored_name = f"{timestamp}_{safe_name}"
            stored_path = os.path.join(self.upload_dir, stored_name)
            shutil.copy2(file_path, stored_path)

            # Extract content
            content = self._extract_content(stored_path, file_type, original_name)

            # Generate summary
            summary = None
            if generate_summary and content:
                summary = self._generate_summary(content[:5000])

            # Chunk for vector search
            chunks = self._chunk_content(content) if content else []

            # Store vector embeddings
            vector_id = None
            if self.encoder and chunks:
                vector_id = self._store_vectors(stored_name, chunks)

            # Save to database
            db.save_document(
                filename=stored_name,
                original_name=original_name,
                file_type=file_type,
                file_size=file_size,
                content=content[:50000] if content else None,  # Limit stored content
                summary=summary,
                chunks=json.dumps(chunks) if chunks else None,
                vector_id=vector_id,
                tags=self._auto_tag(content, original_name) if content else None
            )

            # Log
            db.log("INFO", "document_processor", f"Processed {original_name} ({file_type}, {file_size} bytes)")

            return {
                "success": True,
                "stored_name": stored_name,
                "file_type": file_type,
                "size": file_size,
                "content_preview": content[:500] if content else None,
                "summary": summary,
                "chunks": len(chunks),
                "vector_id": vector_id
            }

        except Exception as e:
            db.log("ERROR", "document_processor", f"Failed to process {original_name}: {e}")
            return {"success": False, "error": str(e)}

    def _extract_content(self, file_path: str, file_type: str, original_name: str) -> str:
        """Extract text content from file"""

        if file_type == 'text' or file_type == 'code':
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                return f.read()

        elif file_type == 'document':
            if original_name.endswith('.pdf'):
                return self._extract_pdf(file_path)
            elif original_name.endswith('.docx'):
                return self._extract_docx(file_path)
            elif original_name.endswith('.doc'):
                return self._extract_doc(file_path)
            else:
                return f"[Document: {original_name} - text extraction not available for this format]"

        elif file_type == 'spreadsheet':
            if original_name.endswith('.xlsx') or original_name.endswith('.xls'):
                return self._extract_excel(file_path)
            else:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()

        elif file_type == 'image':
            return self._extract_image_text(file_path)

        elif file_type == 'archive':
            return self._extract_archive(file_path)

        else:
            return f"[File: {original_name} - content extraction not supported for type: {file_type}]"

    def _extract_pdf(self, file_path: str) -> str:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n\n"
            return text
        except Exception as e:
            return f"[PDF extraction error: {e}]"

    def _extract_docx(self, file_path: str) -> str:
        try:
            from docx import Document
            doc = Document(file_path)
            text = "\n".join([para.text for para in doc.paragraphs])
            return text
        except Exception as e:
            return f"[DOCX extraction error: {e}]"

    def _extract_doc(self, file_path: str) -> str:
        # Legacy .doc format - try antiword or textract
        try:
            result = shutil.which("antiword")
            if result:
                import subprocess
                output = subprocess.run(["antiword", file_path], capture_output=True, text=True, timeout=30)
                return output.stdout
            return "[Legacy .doc format - install antiword for extraction]"
        except Exception as e:
            return f"[DOC extraction error: {e}]"

    def _extract_excel(self, file_path: str) -> str:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            text = []
            for sheet in wb.sheetnames:
                text.append(f"\n=== Sheet: {sheet} ===")
                ws = wb[sheet]
                for row in ws.iter_rows(values_only=True):
                    text.append(" | ".join([str(cell) if cell is not None else "" for cell in row]))
            return "\n".join(text)
        except Exception as e:
            return f"[Excel extraction error: {e}]"

    def _extract_image_text(self, file_path: str) -> str:
        """Extract text from image using OCR"""
        try:
            from PIL import Image
            import pytesseract

            img = Image.open(file_path)
            text = pytesseract.image_to_string(img)

            # Also get image metadata
            metadata = f"Image size: {img.size}, Mode: {img.mode}\n"
            return metadata + "\nExtracted text:\n" + text
        except ImportError:
            return "[Image: OCR not available. Install pytesseract and Tesseract OCR]"
        except Exception as e:
            return f"[Image OCR error: {e}]"

    def _extract_archive(self, file_path: str) -> str:
        """List contents of archive"""
        try:
            import zipfile
            if zipfile.is_zipfile(file_path):
                with zipfile.ZipFile(file_path, 'r') as z:
                    files = z.namelist()
                    return f"ZIP archive containing {len(files)} files:\n" + "\n".join(files[:50])
        except:
            pass
        return "[Archive contents listing not available]"

    def _chunk_content(self, content: str, chunk_size: int = 500, overlap: int = 50) -> List[str]:
        """Split content into overlapping chunks for vector search"""
        if not content:
            return []

        chunks = []
        start = 0
        while start < len(content):
            end = start + chunk_size
            chunk = content[start:end]
            chunks.append(chunk)
            start = end - overlap

        return chunks

    def _store_vectors(self, doc_id: str, chunks: List[str]) -> str:
        """Store vector embeddings for chunks"""
        if not self.encoder:
            return None

        try:
            embeddings = self.encoder.encode(chunks)
            self.vectors[doc_id] = embeddings
            self.documents[doc_id] = chunks
            return doc_id
        except Exception as e:
            db.log("WARNING", "vector_store", f"Failed to store vectors: {e}")
            return None

    def _generate_summary(self, content: str) -> str:
        """Generate document summary using AI"""
        try:
            from aevibron_client import aevibron
            prompt = f"Summarize this document in 2-3 sentences:\n\n{content[:3000]}"
            response = aevibron.fast_response(prompt)
            return response.get("choices", [{}])[0].get("message", {}).get("content", "No summary available")
        except Exception as e:
            return f"Summary generation failed: {e}"

    def _auto_tag(self, content: str, filename: str) -> str:
        """Auto-generate tags based on content"""
        tags = []

        # File extension tag
        ext = Path(filename).suffix.lower()
        if ext:
            tags.append(ext[1:])

        # Content-based tags
        content_lower = content.lower()[:5000]
        keyword_tags = {
            "api": ["api", "endpoint", "rest", "graphql", "swagger"],
            "database": ["database", "sql", "mongodb", "postgres", "mysql", "sqlite"],
            "frontend": ["react", "vue", "angular", "html", "css", "javascript", "typescript"],
            "backend": ["flask", "django", "fastapi", "express", "node", "server"],
            "ai": ["machine learning", "ai", "neural", "model", "training", "llm"],
            "security": ["auth", "password", "token", "encrypt", "security", "jwt"],
            "deployment": ["docker", "kubernetes", "vercel", "aws", "deploy", "ci/cd"]
        }

        for tag, keywords in keyword_tags.items():
            if any(kw in content_lower for kw in keywords):
                tags.append(tag)

        return ",".join(tags)

    def search_documents(self, query: str, semantic: bool = True, limit: int = 10) -> List[Dict]:
        """Search documents by text and/or semantic similarity"""
        results = []

        # Text search via database
        text_results = db.search_documents(query, limit=limit)
        results.extend(text_results)

        # Semantic search via vectors
        if semantic and self.encoder and query:
            try:
                query_vector = self.encoder.encode([query])[0]

                # Simple cosine similarity search
                import numpy as np
                for doc_id, embeddings in self.vectors.items():
                    similarities = np.dot(embeddings, query_vector) / (
                        np.linalg.norm(embeddings, axis=1) * np.linalg.norm(query_vector)
                    )
                    max_sim = np.max(similarities)
                    if max_sim > 0.3:  # Threshold
                        doc = db.get_document(int(doc_id.split('_')[0]) if doc_id.split('_')[0].isdigit() else 1)
                        if doc:
                            doc['semantic_score'] = float(max_sim)
                            results.append(doc)
            except Exception as e:
                db.log("WARNING", "semantic_search", f"Search failed: {e}")

        # Deduplicate and sort
        seen = set()
        unique_results = []
        for r in results:
            doc_id = r.get('id')
            if doc_id not in seen:
                seen.add(doc_id)
                unique_results.append(r)

        return unique_results[:limit]

    def analyze_document(self, doc_id: int) -> Dict:
        """Deep analysis of a document"""
        doc = db.get_document(doc_id)
        if not doc:
            return {"success": False, "error": "Document not found"}

        content = doc.get('content', '')

        analysis = {
            "word_count": len(content.split()) if content else 0,
            "line_count": content.count('\n') if content else 0,
            "file_type": doc.get('file_type'),
            "size": doc.get('file_size'),
            "tags": doc.get('tags'),
            "summary": doc.get('summary'),
            "key_findings": []
        }

        # Extract key findings
        if content:
            # Find URLs
            urls = re.findall(r'https?://[^\s"'<>]+', content)
            if urls:
                analysis["key_findings"].append(f"Found {len(urls)} URLs")

            # Find code blocks
            code_blocks = re.findall(r'```[\s\S]*?```', content)
            if code_blocks:
                analysis["key_findings"].append(f"Found {len(code_blocks)} code blocks")

            # Find TODOs
            todos = re.findall(r'TODO|FIXME|HACK|BUG|XXX', content, re.IGNORECASE)
            if todos:
                analysis["key_findings"].append(f"Found {len(todos)} TODOs/FIXMEs")

        return {"success": True, "analysis": analysis}

    def get_document_stats(self) -> Dict:
        """Get statistics about all documents"""
        conn = db.get_connection()
        cursor = conn.execute("SELECT COUNT(*) as total FROM documents")
        total = cursor.fetchone()[0]

        cursor = conn.execute("SELECT file_type, COUNT(*) as count FROM documents GROUP BY file_type")
        by_type = {r['file_type']: r['count'] for r in cursor.fetchall()}

        cursor = conn.execute("SELECT SUM(file_size) as total_size FROM documents")
        total_size = cursor.fetchone()[0] or 0

        conn.close()

        return {
            "total_documents": total,
            "by_type": by_type,
            "total_size_mb": round(total_size / (1024*1024), 2),
            "vector_store_size": len(self.vectors)
        }

# Singleton
document_processor = DocumentProcessor()
